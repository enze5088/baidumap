# coding=utf-8
from urllib import request
import time
import datetime
import random
import CoreMapApi
from openpyxl import Workbook
from tqdm import tqdm



def write_migrate_book(ws,result,column,row,Date):
    '''

    :param ws:
    :param result:
    :param column:
    :param row:
    :return:

    '''
    ws.cell(row=row, column=column+1, value='city')
    ws.cell(row=row, column=column+2, value='province_name')
    ws.cell(row=row, column=column+3, value=str(Date.strftime('%Y%m%d')))
    row = row +1
    try:
        for ob in result:
            city_name = ob['city_name']
            province_name = ob['province_name']
            value = ob['value']
            ws.cell(row=row, column=column+1, value=city_name)
            ws.cell(row=row, column=column+2, value=province_name)
            ws.cell(row=row, column=column+3, value=value)
            row = row + 1
    except:
        print('该城市数据写入错误')


if __name__ == '__main__':

    wb = Workbook()

    headers = {
        "User-agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.221 Safari/537.36 SE 2.X MetaSr 1.0"}
    opener = request.build_opener()
    opener.add_headers = [headers]
    request.install_opener(opener)
    endDate = datetime.datetime.now()+datetime.timedelta(days=-1)
    #types = ["move_in","move_out"]
    types = ["move_out"]
    index = 0
    city = "武汉"
    citys = []
    for item in CoreMapApi.AllName.values():
        citys.extend(item)
    dts = ["city",'province']
    T = 14
    Date = datetime.datetime.strptime('20200110', '%Y%m%d')
    DateList = [Date + datetime.timedelta(n) for n in range(T + 1)]
    for type in types:
        for city in tqdm(citys):
            column = 0
            row = 1
            ws = wb.create_sheet(title=city)
            for D in DateList:
                dt ='city'
                try:
                    result = CoreMapApi.MigratePost(dt,city,type,D)
                    write_migrate_book(ws=ws,result=result,column=column,row=row,Date=D)
                except:
                    print(city)
                column = column + 3
                time.sleep(random.uniform(0.08,0.3))
        print("完成")
        wb.remove(wb['Sheet'])
        dest_filename = './Data/migrate/'+type+'migrate_book.xlsx'
        wb.save(dest_filename)

