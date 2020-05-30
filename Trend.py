# coding=utf-8
import requests
from urllib import request
import time
import datetime
import random
import json
import openpyxl
from openpyxl import Workbook
import CoreMapApi
import collections





def write_in_book(ws,result,column,row,name,type):
    '''

    :param ws:
    :param result:
    :param column:
    :param row:
    :return:
    '''
    ws.cell(row=row, column=column, value=name)
    column = column + 1

    ws.cell(row=row, column=column, value=type)
    column = column + 1

    for key,value in result.items():
        ws.cell(row=row,column=column,value=value)
        ws.cell(row=1,column=column,value=key)
        column = column+1


if __name__ == '__main__':
    dest_filename = './Data/migrate_Trend.xlsx'
    wb = Workbook()
    headers = {
        "User-agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.221 Safari/537.36 SE 2.X MetaSr 1.0"}
    opener = request.build_opener()
    opener.add_headers = [headers]
    request.install_opener(opener)
    endDate = datetime.datetime.now()+datetime.timedelta(days=-1)
    types = ["move_in","move_out","internalflowhistory"]
    index = 0

    for type in types:
        column = 1
        row = 2
        ws = wb.create_sheet(title=type)
        for key,value in CoreMapApi.AllName.items():
            if key == "直辖市":
                dt = 'province'
                for city in value:
                    result = CoreMapApi.TrendPost(dt,city,type,endDate)
                    write_in_book(ws=ws,result=result,column=column,row=row,name=city,type=key)
                    row = row + 1
                    time.sleep(random.uniform(0.08,0.3))

            elif not type =="internalflowhistory":
                dt = 'province'
                city = key
                result = CoreMapApi.TrendPost(dt, city, type, endDate)
                write_in_book(ws=ws, result=result, column=column, row=row, name=city, type="省份")
                row = row + 1
                time.sleep(random.uniform(0.1, 0.4))
                for city in value:
                    dt = 'city'
                    result = CoreMapApi.TrendPost(dt, city, type, endDate)
                    write_in_book(ws=ws, result=result, column=column, row=row, name=city, type="城市")
                    row = row + 1
                    time.sleep(random.uniform(0.08, 0.3))

    print("完成")
    wb.remove(wb['Sheet'])
    wb.save(dest_filename)

